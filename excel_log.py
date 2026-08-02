"""Logs music folders added to the library into an Excel (.xlsx)
workbook, one row per track discovered (Album / Artist / Year columns --
mirroring the layout of a manually kept "disc organizer" spreadsheet),
grouped under a small section header noting which folder it came from
and when it was added.

Uses openpyxl (a pure-Python .xlsx reader/writer) rather than anything
LibreOffice/Excel-specific, so no external office suite needs to be
installed to create or update the log file.

See App._log_scan_to_excel / App._ensure_library_log_path in app.py for
how/when this gets called.
"""

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

SHEET_NAME = "Library Log"
COLUMN_HEADERS = ["Album", "Artist", "Year"]


def append_folder_log(xlsx_path, folder_path, entries):
    """Append a new section to `xlsx_path` for `folder_path`: a small
    header row (folder path + timestamp), a column-header row, then one
    row per (album, artist, year) tuple in `entries`, followed by a
    blank spacer row. Creates the workbook (with a `SHEET_NAME` sheet)
    if `xlsx_path` doesn't exist yet, or appends to it (adding the sheet
    too, if missing) if it does. Raises on failure (e.g. the file is
    open/locked elsewhere, or the path isn't writable) -- callers should
    catch and report that themselves."""
    if os.path.isfile(xlsx_path):
        workbook = load_workbook(xlsx_path)
        if SHEET_NAME in workbook.sheetnames:
            sheet = workbook[SHEET_NAME]
        else:
            sheet = workbook.create_sheet(SHEET_NAME)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([f"Folder: {folder_path}", f"Added: {timestamp}"])
    sheet.append(list(COLUMN_HEADERS))
    for album, artist, year in entries:
        sheet.append([album, artist, year])
    # Note: deliberately no trailing "blank spacer row" here -- a fully
    # empty row (no cells with any value) isn't reliably preserved by
    # openpyxl across a save + reload cycle (each subsequent folder's
    # section reopens/re-saves the file), so it would just silently
    # disappear again next time this is called. The "Folder: ..." header
    # row above already makes each section's start clear enough without it.

    parent_dir = os.path.dirname(xlsx_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    workbook.save(xlsx_path)
